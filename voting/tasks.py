from __future__ import absolute_import, unicode_literals
from io import BytesIO
from src.celery import app
from voting.models import Voting
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from django.core.files import File
from django.core.mail import EmailMessage


@app.task
def create_xlsx(voting_id):
    print('я тут')
    voting = Voting.objects.get(id=voting_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Результаты голосования'

    person_votes = voting.person_votes.all()

    ws.append(('Персонажи', 'Голоса'))
    for person_vote in person_votes:
        row = (str(person_vote.person.get_full_name()), int(person_vote.votes))
        ws.append(row)

    ws.column_dimensions[get_column_letter(1)].width = 40

    chart = BarChart()
    chart.type = 'bar'
    chart.style = 10
    chart.title = str(voting.title)
    chart.x_axis.title = 'Персонажи'
    chart.y_axis.title = 'Голоса'

    chart.y_axis.scaling.min = 0
    if voting.max_votes:
        chart.y_axis.scaling.max = int(voting.max_votes)

    max_row = len(person_votes) + 1

    data = Reference(ws, min_col=2, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, f'A{max_row + 3}')

    io = BytesIO(save_virtual_workbook(wb))

    voting.xlsx.save(f'{voting.title}.xlsx', File(io))
    voting.xlsx_status = 1
    voting.save()


@app.task
def send_email(email):
    email = EmailMessage('Файл сгененрирован', 'Файл сгенерирован', to=(email,))
    email.send()
